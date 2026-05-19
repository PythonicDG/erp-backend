from rest_framework import viewsets, filters, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import Project, CustomerMaster, StandardMaster, InspectionAuthorityMaster, ECN
from .serializers import ProjectSerializer, CustomerMasterSerializer, StandardMasterSerializer, InspectionAuthorityMasterSerializer, ECNSerializer
from .permissions import CanCreateProject
from authentication.mixins import AuditLogMixin
from authentication.permissions import IsAdmin
from django.db.models import Q

class ProjectViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Project.objects.all().select_related('customer').prefetch_related(
        'workflow_stages', 
        'workflow_stages__template'
    )
    serializer_class = ProjectSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = ['status', 'project_type', 'customer_name']
    search_fields = ['pid', 'name', 'customer_name', 'customer_part_no', 'pcepl_part_no']
    ordering_fields = ['date_received', 'target_completion_date', 'created_at', 'pid']
    ordering = ['-created_at']
    audit_module = "Projects"

    def get_audit_target(self, instance):
        return f"Project: {instance.pid} ({instance.name})"

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
        super().perform_create(serializer)

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects Template"
        
        # Headers
        headers = [
            "Project Name", "Customer Name", "Customer Part Number", 
            "PCEPL Part Number", "Inspection Authority", "Applicable Standard", 
            "Date Received", "Target Completion Date", "Status", "Priority", 
            "Assigned To", "Description"
        ]
        ws.append(headers)
        
        # Sample Data
        sample_rows = [
            [
                "Junction Box CSL-WATERJET", "Larsen & Toubro", "PAAG464305", 
                "30071850", "Internal QA", "IEC 61439", "2026-05-17", 
                "2026-06-30", "Open", "High", "employee@example.com", 
                "Standard Junction Box assembly and wiring."
            ],
            [
                "CSL Panel Maintenance", "Cochin Shipyard", "CSL-9988", 
                "", "LRS", "IEC 60947", "2026-05-20", 
                "", "Draft", "Medium", "", "Annual maintenance and QA inspection."
            ]
        ]
        for row in sample_rows:
            ws.append(row)
            
        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="projects_bulk_upload_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Apply standard queryset filtering (search & filters)
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Master Export"
        
        # Premium Styling
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep navy
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        headers = [
            "Project ID (PID)", "Project Name", "Customer Name", "Customer Part Number", 
            "PCEPL Part Number", "Inspection Authority", "Applicable Standard", 
            "Date Received", "Target Completion Date", "Status", "Priority", 
            "Assigned Employee", "Supervisor", "Created By", "Date Created"
        ]
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Write data
        for project in queryset:
            row = [
                project.pid,
                project.name,
                project.customer.name if project.customer else project.customer_name or '',
                project.customer_part_no or '',
                project.pcepl_part_no or '',
                project.inspection_authority or '',
                project.applicable_standard or '',
                project.date_received.strftime('%Y-%m-%d') if project.date_received else '',
                project.target_completion_date.strftime('%Y-%m-%d') if project.target_completion_date else '',
                project.status,
                project.priority or '',
                project.assigned_employee.full_name if project.assigned_employee else '',
                project.supervisor.full_name if project.supervisor else '',
                project.created_by.full_name if project.created_by else '',
                project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
            ]
            ws.append(row)
            
        # Format data cells
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [1, 8, 9, 10, 11, 15]: # PID, Dates, Status, Priority
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"Project_Master_Export_{current_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        # Log action for audit logs
        from authentication.utils import log_action
        log_action(
            user=request.user,
            action="EXPORT",
            target="All Projects Excel",
            module="Projects"
        )
        
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        import openpyxl
        from datetime import datetime, date, timedelta
        from django.db import transaction
        from django.contrib.auth import get_user_model
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
            
        skip_duplicates = request.data.get('skip_duplicates', 'true').lower() == 'true'
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            
            name_idx = -1
            client_idx = -1
            cust_part_idx = -1
            pcepl_part_idx = -1
            inspect_auth_idx = -1
            standard_idx = -1
            start_idx = -1
            end_idx = -1
            status_idx = -1
            priority_idx = -1
            assigned_idx = -1
            desc_idx = -1

            for idx, h in enumerate(headers):
                if h in ['project name', 'project_name', 'name', 'project']:
                    name_idx = idx
                elif h in ['customer name', 'customer_name', 'client name', 'client_name', 'client', 'customer']:
                    client_idx = idx
                elif h in ['customer part number', 'customer_part_number', 'customer part no', 'customer_part_no']:
                    cust_part_idx = idx
                elif h in ['pcepl part number', 'pcepl_part_number', 'pcepl part no', 'pcepl_part_no']:
                    pcepl_part_idx = idx
                elif h in ['inspection authority', 'inspection_authority', 'inspection authority/agency']:
                    inspect_auth_idx = idx
                elif h in ['applicable standard', 'applicable_standard', 'standard']:
                    standard_idx = idx
                elif h in ['date received', 'date_received', 'start date', 'start_date', 'start']:
                    start_idx = idx
                elif h in ['target completion date', 'target_completion_date', 'end date', 'end_date', 'end']:
                    end_idx = idx
                elif h in ['status']:
                    status_idx = idx
                elif h in ['priority']:
                    priority_idx = idx
                elif h in ['assigned to', 'assigned_to', 'assigned employee', 'assigned_employee', 'assigned']:
                    assigned_idx = idx
                elif h in ['description', 'desc', 'remarks', 'notes']:
                    desc_idx = idx
                    
            missing = []
            if name_idx == -1: missing.append("Project Name")
            if client_idx == -1: missing.append("Customer Name")
            if start_idx == -1: missing.append("Date Received")

            if missing:
                return Response({
                    "error": f"Missing required columns: {', '.join(missing)}. Please use the downloadable template."
                }, status=status.HTTP_400_BAD_REQUEST)
                
            def parse_date(val):
                if not val:
                    return None
                if isinstance(val, (date, datetime)):
                    return val.date() if isinstance(val, datetime) else val
                # If string, try multiple formats
                val_str = str(val).strip()
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(val_str, fmt).date()
                    except ValueError:
                        continue
                # Try parsing Excel float serial date
                try:
                    if val_str.replace('.', '', 1).isdigit():
                        serial = float(val_str)
                        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
                except Exception:
                    pass
                raise ValueError(f"Invalid date format: {val}")

            success_count = 0
            skipped_count = 0
            failure_count = 0
            errors = []
            skipped = []
            successes = []
            User = get_user_model()
            row_num = 1 # fallback

            for row_num, row in enumerate(rows[1:], start=2):
                # Check if row is empty
                if not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue
                    
                proj_name = ""
                client_name = ""
                try:
                    proj_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                    client_name = str(row[client_idx]).strip() if row[client_idx] is not None else ""
                    
                    if not proj_name:
                        errors.append({
                            "row": row_num,
                            "project_name": "N/A",
                            "error_message": "Project Name is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not client_name:
                        errors.append({
                            "row": row_num,
                            "project_name": proj_name,
                            "error_message": "Customer Name is required"
                        })
                        failure_count += 1
                        continue
                        
                    start_date_val = row[start_idx]
                    if start_date_val is None or str(start_date_val).strip() == "":
                        errors.append({
                            "row": row_num,
                            "project_name": proj_name,
                            "error_message": "Date Received is required"
                        })
                        failure_count += 1
                        continue
                        
                    try:
                        date_received = parse_date(start_date_val)
                    except ValueError as ve:
                        errors.append({
                            "row": row_num,
                            "project_name": proj_name,
                            "error_message": str(ve)
                        })
                        failure_count += 1
                        continue
                        
                    target_completion_date = None
                    if end_idx != -1 and row[end_idx] is not None and str(row[end_idx]).strip() != "":
                        try:
                            target_completion_date = parse_date(row[end_idx])
                        except ValueError as ve:
                            errors.append({
                                "row": row_num,
                                "project_name": proj_name,
                                "error_message": f"Target Completion Date error: {str(ve)}"
                            })
                            failure_count += 1
                            continue
                            
                    status_val = "Open"
                    if status_idx != -1 and row[status_idx] is not None:
                        status_str = str(row[status_idx]).strip()
                        from .models import ProjectStatus
                        matched = False
                        for choice_val, choice_label in ProjectStatus.choices:
                            if status_str.lower() == choice_val.lower() or status_str.lower() == choice_label.lower():
                                status_val = choice_val
                                matched = True
                                break
                        if not matched:
                            status_val = "Open"
                            
                    priority_val = ""
                    if priority_idx != -1 and row[priority_idx] is not None:
                        priority_val = str(row[priority_idx]).strip()
                        
                    desc_val = ""
                    if desc_idx != -1 and row[desc_idx] is not None:
                        desc_val = str(row[desc_idx]).strip()
                        
                    cust_part_val = ""
                    if cust_part_idx != -1 and row[cust_part_idx] is not None:
                        cust_part_val = str(row[cust_part_idx]).strip()
                        
                    pcepl_part_val = ""
                    if pcepl_part_idx != -1 and row[pcepl_part_idx] is not None:
                        pcepl_part_val = str(row[pcepl_part_idx]).strip()
                        
                    inspect_auth_val = ""
                    if inspect_auth_idx != -1 and row[inspect_auth_idx] is not None:
                        inspect_auth_val = str(row[inspect_auth_idx]).strip()
                        
                    standard_val = ""
                    if standard_idx != -1 and row[standard_idx] is not None:
                        standard_val = str(row[standard_idx]).strip()

                    customer = CustomerMaster.objects.filter(name__iexact=client_name).first()
                    if not customer:
                        customer = CustomerMaster.objects.create(
                            name=client_name,
                            email=f"{client_name.lower().replace(' ', '')}@example.com",
                            mobile_number="0000000000",
                            remarks="Auto-created during bulk upload"
                        )
                        
                    if skip_duplicates:
                        existing_project = Project.objects.filter(name__iexact=proj_name, customer=customer).first()
                        if existing_project:
                            skipped.append({
                                "row": row_num,
                                "project_name": proj_name,
                                "reason": f"Project '{proj_name}' already exists for client '{client_name}'"
                            })
                            skipped_count += 1
                            continue
                            
                    assigned_employee = None
                    if assigned_idx != -1 and row[assigned_idx] is not None:
                        assigned_to_str = str(row[assigned_idx]).strip()
                        if assigned_to_str:
                            if '@' in assigned_to_str:
                                assigned_employee = User.objects.filter(email__iexact=assigned_to_str).first()
                            else:
                                parts = assigned_to_str.split(' ', 1)
                                if len(parts) == 2:
                                    assigned_employee = User.objects.filter(
                                        first_name__iexact=parts[0],
                                        last_name__iexact=parts[1]
                                    ).first()
                                if not assigned_employee:
                                    assigned_employee = User.objects.filter(
                                        first_name__iexact=assigned_to_str
                                    ).first()
                                    
                    with transaction.atomic():
                        project = Project.objects.create(
                            name=proj_name,
                            customer=customer,
                            customer_name=customer.name,
                            customer_part_no=cust_part_val,
                            pcepl_part_no=pcepl_part_val,
                            inspection_authority=inspect_auth_val,
                            applicable_standard=standard_val,
                            date_received=date_received,
                            target_completion_date=target_completion_date,
                            status=status_val,
                            priority=priority_val,
                            description=desc_val,
                            assigned_employee=assigned_employee,
                            created_by=request.user,
                            project_type="OTHER"
                        )
                        
                        from .services import ProjectService
                        ProjectService.initialize_workflow(project)
                        
                        ProjectService.log_activity(
                            project,
                            request.user,
                            "Project Created via Bulk Upload",
                            {"status": project.status}
                        )
                        
                    successes.append(proj_name)
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": row_num,
                        "project_name": proj_name if proj_name else "Unknown",
                        "error_message": f"Unexpected error: {str(e)}"
                    })
                    failure_count += 1
                    
            return Response({
                "success_count": success_count,
                "skipped_count": skipped_count,
                "failure_count": failure_count,
                "total_processed": len(rows) - 1,
                "errors": errors,
                "skipped": skipped,
                "successes": successes
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

class CustomerMasterViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = CustomerMaster.objects.all()
    serializer_class = CustomerMasterSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_fields = ['category']
    search_fields = ['name', 'mobile_number', 'email']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    permission_classes = [permissions.IsAuthenticated]
    audit_module = "Customer Masters"

    def get_audit_target(self, instance):
        return f"Customer: {instance.name}"

    def get_permissions(self):
        if self.action == 'create':
            return [CanCreateProject()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Apply filters & search context
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Master Export"
        
        # Professional Colors & Styles
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Navy Blue
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        headers = [
            "Customer Name", "Customer Category", "Mobile Number", 
            "Alternate Mobile Number", "Email Address", "Remarks / Notes", 
            "Date Created", "Last Updated"
        ]
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Write customer data
        for customer in queryset:
            row = [
                customer.name,
                customer.category or '',
                customer.mobile_number or '',
                customer.alternate_mobile_number or '',
                customer.email or '',
                customer.remarks or '',
                customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else '',
                customer.updated_at.strftime('%Y-%m-%d %H:%M:%S') if customer.updated_at else ''
            ]
            ws.append(row)
            
        # Format cells
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [2, 3, 4, 7, 8]: # Category, Phone numbers, Dates
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"Customer_Master_Export_{current_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        # Audit Log support
        from authentication.utils import log_action
        log_action(
            user=request.user,
            action="EXPORT",
            target="All Customers Excel",
            module="Customer Masters"
        )
        
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers Template"
        
        # Headers
        headers = [
            "Customer Name", "Email Address", "Category", 
            "Mobile Number", "Alternate Mobile", "Remarks / Notes"
        ]
        ws.append(headers)
        
        # Sample Data
        sample_rows = [
            [
                "Larsen & Toubro", "larsen.toubro@example.com", "A-Category", 
                "9876543210", "0222456789", "Key client for heavy electrical junction boxes."
            ],
            [
                "Cochin Shipyard", "cochin.shipyard@example.com", "B-Category", 
                "8887776665", "", "Handles naval projects panel supply."
            ]
        ]
        for row in sample_rows:
            ws.append(row)
            
        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="customers_bulk_upload_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        import openpyxl
        from django.db import transaction
        from authentication.system_models import AuditLog
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
            
        skip_duplicates = request.data.get('skip_duplicates', 'true').lower() == 'true'
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            
            name_idx = -1
            email_idx = -1
            cat_idx = -1
            mobile_idx = -1
            alt_mobile_idx = -1
            remarks_idx = -1

            for idx, h in enumerate(headers):
                if h in ['customer name', 'customer_name', 'name', 'customer']:
                    name_idx = idx
                elif h in ['email address', 'email_address', 'email']:
                    email_idx = idx
                elif h in ['category', 'customer category', 'customer_category']:
                    cat_idx = idx
                elif h in ['mobile number', 'mobile_number', 'mobile', 'phone']:
                    mobile_idx = idx
                elif h in ['alternate mobile', 'alternate_mobile', 'alternate mobile number', 'alternate_mobile_number', 'alt mobile', 'alt_mobile']:
                    alt_mobile_idx = idx
                elif h in ['remarks / notes', 'remarks', 'notes', 'remarks_notes', 'remarks/notes', 'remark']:
                    remarks_idx = idx
                    
            missing = []
            if name_idx == -1: missing.append("Customer Name")
            if email_idx == -1: missing.append("Email Address")
            if mobile_idx == -1: missing.append("Mobile Number")

            if missing:
                return Response({
                    "error": f"Missing required columns: {', '.join(missing)}. Please use the downloadable template."
                }, status=status.HTTP_400_BAD_REQUEST)

            success_count = 0
            skipped_count = 0
            failure_count = 0
            errors = []
            skipped = []
            successes = []

            for row_num, row in enumerate(rows[1:], start=2):
                # Check if row is empty
                if not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue
                    
                cust_name = ""
                try:
                    cust_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                    email_val = str(row[email_idx]).strip() if row[email_idx] is not None else ""
                    mobile_val = str(row[mobile_idx]).strip() if row[mobile_idx] is not None else ""
                    
                    if not cust_name:
                        errors.append({
                            "row": row_num,
                            "customer_name": "N/A",
                            "error_message": "Customer Name is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not email_val:
                        errors.append({
                            "row": row_num,
                            "customer_name": cust_name,
                            "error_message": "Email Address is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not mobile_val:
                        errors.append({
                            "row": row_num,
                            "customer_name": cust_name,
                            "error_message": "Mobile Number is required"
                        })
                        failure_count += 1
                        continue

                    category_val = ""
                    if cat_idx != -1 and row[cat_idx] is not None:
                        category_val = str(row[cat_idx]).strip()
                        
                    alt_mobile_val = ""
                    if alt_mobile_idx != -1 and row[alt_mobile_idx] is not None:
                        alt_mobile_val = str(row[alt_mobile_idx]).strip()
                        
                    remarks_val = ""
                    if remarks_idx != -1 and row[remarks_idx] is not None:
                        remarks_val = str(row[remarks_idx]).strip()
                        
                    if skip_duplicates:
                        existing_customer = CustomerMaster.objects.filter(
                            Q(name__iexact=cust_name) | Q(email__iexact=email_val)
                        ).first()
                        if existing_customer:
                            skipped.append({
                                "row": row_num,
                                "customer_name": cust_name,
                                "reason": f"Customer with name '{cust_name}' or email '{email_val}' already exists."
                            })
                            skipped_count += 1
                            continue
                            
                    with transaction.atomic():
                        customer = CustomerMaster.objects.create(
                            name=cust_name,
                            email=email_val,
                            category=category_val,
                            mobile_number=mobile_val,
                            alternate_mobile_number=alt_mobile_val,
                            remarks=remarks_val
                        )
                        
                        AuditLog.objects.create(
                            user=request.user,
                            action="Customer Created via Bulk Upload",
                            target=f"Customer: {customer.name}",
                            module="Customer Masters",
                            status="SUCCESS"
                        )
                        
                    successes.append(cust_name)
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": row_num,
                        "customer_name": cust_name if cust_name else "Unknown",
                        "error_message": f"Unexpected error: {str(e)}"
                    })
                    failure_count += 1
                    
            return Response({
                "success_count": success_count,
                "skipped_count": skipped_count,
                "failure_count": failure_count,
                "total_processed": len(rows) - 1,
                "errors": errors,
                "skipped": skipped,
                "successes": successes
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def sync_all_statuses(self, request):
        from workflow.models import StageInstance
        projects = Project.objects.exclude(status='Closed')
        updated_count = 0
        for project in projects:
            instances = StageInstance.objects.filter(project=project)
            if instances.exists() and all(inst.status == 'Approved' for inst in instances):
                project.status = 'Closed'
                project.save()
                updated_count += 1
        return Response({"message": f"Updated {updated_count} projects"})

    @action(detail=True, methods=['get'])
    def full_report(self, request, pk=None):
        from workflow.models import StageInstance
        from workflow.serializers import StageInstanceSerializer
        from authentication.system_models import CompanyProfile
        from authentication.serializers import CompanyProfileSerializer
        
        project = self.get_object()
        stages = StageInstance.objects.filter(project=project).order_by('order')
        company = CompanyProfile.objects.first()
        
        return Response({
            "project": ProjectSerializer(project).data,
            "stages": StageInstanceSerializer(stages, many=True).data,
            "company": CompanyProfileSerializer(company).data if company else None
        })

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        from django.db.models import Count
        stats = Project.objects.values('status').annotate(count=Count('id'))
        return Response(stats)


class StandardMasterViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = StandardMaster.objects.all()
    serializer_class = StandardMasterSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    filterset_fields = ['category', 'status', 'release_year']
    search_fields = ['standard_number', 'standard_name']
    ordering_fields = ['standard_number', 'release_year', 'created_at']
    ordering = ['standard_number']
    permission_classes = [permissions.IsAuthenticated]
    audit_module = "Standards Master"

    def get_audit_target(self, instance):
        return f"Standard: {instance.standard_number} - {instance.standard_name}"

    def get_permissions(self):
        # Allow safe GET actions for all authenticated users to read active list (e.g., for dropdown selection)
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        # Full CRUD/Bulk operations restricted to ADMIN role only
        from .permissions import IsAdminUser
        return [permissions.IsAuthenticated(), IsAdminUser()]

    def get_queryset(self):
        user = self.request.user
        queryset = StandardMaster.objects.all()
        # If user is not Admin, restrict lists to Active status only
        if user and user.role != 'ADMIN':
            queryset = queryset.filter(status='Active')
        return queryset

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Standards Master Export"
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        headers = [
            "Standard Number", "Standard Name", "Revision / Edition", 
            "Release Year", "Category", "Description", "Status", 
            "Date Created", "Last Updated"
        ]
        
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        for std in queryset:
            row = [
                std.standard_number,
                std.standard_name,
                std.revision or '',
                std.release_year or '',
                std.category,
                std.description or '',
                std.status,
                std.created_at.strftime('%Y-%m-%d %H:%M:%S') if std.created_at else '',
                std.updated_at.strftime('%Y-%m-%d %H:%M:%S') if std.updated_at else ''
            ]
            ws.append(row)
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                if col_idx in [3, 4, 5, 7, 8, 9]: # Revision, Year, Category, Status, Dates
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"Standard_Master_Export_{current_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        from authentication.utils import log_action
        log_action(
            user=request.user,
            action="EXPORT",
            target="All Standards Excel",
            module="Standards Master"
        )
        
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Standards Template"
        
        headers = [
            "Standard Number", "Standard Name", "Revision / Edition", 
            "Release Year", "Category", "Description", "Status"
        ]
        ws.append(headers)
        
        sample_rows = [
            ["IEC 61439-1", "Low-voltage switchgear and controlgear assemblies", "Edition 3.0", 2020, "IEC", "General rules and guidelines for low-voltage switchboard designs.", "Active"],
            ["ISO 9001", "Quality management systems - Requirements", "Fifth Edition", 2015, "ISO", "Standard for quality management guidelines and execution.", "Active"],
            ["MIL-STD-810G", "Environmental Engineering Considerations and Laboratory Tests", "Rev G", 2008, "Defence", "Military standard for environmental conditions testing.", "Active"]
        ]
        for row in sample_rows:
            ws.append(row)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="standards_bulk_upload_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        import openpyxl
        from django.db import transaction
        from authentication.system_models import AuditLog
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
            
        skip_duplicates = request.data.get('skip_duplicates', 'true').lower() == 'true'
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            
            num_idx = -1
            name_idx = -1
            rev_idx = -1
            year_idx = -1
            cat_idx = -1
            desc_idx = -1
            status_idx = -1

            for idx, h in enumerate(headers):
                if h in ['standard number', 'standard_number', 'number']:
                    num_idx = idx
                elif h in ['standard name', 'standard_name', 'name']:
                    name_idx = idx
                elif h in ['revision', 'edition', 'revision / edition', 'revision_edition']:
                    rev_idx = idx
                elif h in ['release year', 'release_year', 'year']:
                    year_idx = idx
                elif h in ['category', 'standard category', 'standard_category']:
                    cat_idx = idx
                elif h in ['description', 'notes', 'remarks']:
                    desc_idx = idx
                elif h in ['status', 'active', 'state']:
                    status_idx = idx
                    
            missing = []
            if num_idx == -1: missing.append("Standard Number")
            if name_idx == -1: missing.append("Standard Name")
            if cat_idx == -1: missing.append("Category")

            if missing:
                return Response({
                    "error": f"Missing required columns: {', '.join(missing)}. Please use the downloadable template."
                }, status=status.HTTP_400_BAD_REQUEST)

            success_count = 0
            skipped_count = 0
            failure_count = 0
            errors = []
            skipped = []
            successes = []

            # Allowed categories list to validate Category
            allowed_cats = ['ISO', 'IEC', 'Marine IEC', 'IP', 'EMC', 'Defence']

            for row_num, row in enumerate(rows[1:], start=2):
                if not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue
                    
                std_num = ""
                try:
                    std_num = str(row[num_idx]).strip() if row[num_idx] is not None else ""
                    std_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                    cat_val = str(row[cat_idx]).strip() if row[cat_idx] is not None else ""
                    
                    if not std_num:
                        errors.append({
                            "row": row_num,
                            "standard_number": "N/A",
                            "error_message": "Standard Number is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not std_name:
                        errors.append({
                            "row": row_num,
                            "standard_number": std_num,
                            "error_message": "Standard Name is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not cat_val:
                        errors.append({
                            "row": row_num,
                            "standard_number": std_num,
                            "error_message": "Category is required"
                        })
                        failure_count += 1
                        continue

                    # Validate category choice
                    match_cat = next((c for c in allowed_cats if c.lower() == cat_val.lower()), None)
                    if not match_cat:
                        errors.append({
                            "row": row_num,
                            "standard_number": std_num,
                            "error_message": f"Invalid category '{cat_val}'. Must be one of: {', '.join(allowed_cats)}"
                        })
                        failure_count += 1
                        continue

                    rev_val = str(row[rev_idx]).strip() if (rev_idx != -1 and row[rev_idx] is not None) else ""
                    
                    year_val = None
                    if year_idx != -1 and row[year_idx] is not None:
                        try:
                            year_val = int(float(str(row[year_idx]).strip()))
                        except:
                            errors.append({
                                "row": row_num,
                                "standard_number": std_num,
                                "error_message": f"Invalid year value '{row[year_idx]}'. Must be a whole number."
                            })
                            failure_count += 1
                            continue

                    desc_val = str(row[desc_idx]).strip() if (desc_idx != -1 and row[desc_idx] is not None) else ""
                    
                    status_val = "Active"
                    if status_idx != -1 and row[status_idx] is not None:
                        raw_status = str(row[status_idx]).strip().capitalize()
                        if raw_status in ["Active", "Inactive"]:
                            status_val = raw_status

                    if skip_duplicates:
                        existing = StandardMaster.objects.filter(standard_number__iexact=std_num).first()
                        if existing:
                            skipped.append({
                                "row": row_num,
                                "standard_number": std_num,
                                "reason": f"Standard with number '{std_num}' already exists."
                            })
                            skipped_count += 1
                            continue
                            
                    with transaction.atomic():
                        standard_instance, created = StandardMaster.objects.update_or_create(
                            standard_number=std_num,
                            defaults={
                                "standard_name": std_name,
                                "revision": rev_val,
                                "release_year": year_val,
                                "category": match_cat,
                                "description": desc_val,
                                "status": status_val
                            }
                        )
                        
                        AuditLog.objects.create(
                            user=request.user,
                            action="Standard Created/Updated via Bulk Upload",
                            target=f"Standard: {standard_instance.standard_number}",
                            module="Standards Master",
                            status="SUCCESS"
                        )
                        
                    successes.append(std_num)
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": row_num,
                        "standard_number": std_num if std_num else "Unknown",
                        "error_message": f"Unexpected error: {str(e)}"
                    })
                    failure_count += 1
                    
            return Response({
                "success_count": success_count,
                "skipped_count": skipped_count,
                "failure_count": failure_count,
                "total_processed": len(rows) - 1,
                "errors": errors,
                "skipped": skipped,
                "successes": successes
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class InspectionAuthorityMasterViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = InspectionAuthorityMaster.objects.all().select_related('applicable_standard')
    serializer_class = InspectionAuthorityMasterSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'category', 'approval_type']
    search_fields = ['authority_id', 'name', 'contact_person', 'remarks']
    ordering_fields = ['authority_id', 'name', 'created_at']
    ordering = ['authority_id']
    audit_module = "Inspection Authority Master"

    def get_audit_target(self, instance):
        return f"Inspection Authority: {instance.authority_id} ({instance.name})"

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        from .permissions import IsAdminUser
        return [permissions.IsAuthenticated(), IsAdminUser()]

    def get_queryset(self):
        user = self.request.user
        queryset = InspectionAuthorityMaster.objects.all().select_related('applicable_standard')
        if user and user.role != 'ADMIN':
            queryset = queryset.filter(status='Active')
        return queryset

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Authorities Export"
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        headers = [
            "Inspection Authority ID", "Inspection Authority Name", "Category", 
            "Contact Person", "Applicable Standard / Agency", "Approval Type", 
            "Status", "Remarks", "Date Created", "Last Updated"
        ]
        
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        for auth in queryset:
            std_str = f"{auth.applicable_standard.standard_number} - {auth.applicable_standard.standard_name}" if auth.applicable_standard else ''
            row = [
                auth.authority_id,
                auth.name,
                auth.category,
                auth.contact_person or '',
                std_str,
                auth.approval_type or '',
                auth.status,
                auth.remarks or '',
                auth.created_at.strftime('%Y-%m-%d %H:%M:%S') if auth.created_at else '',
                auth.updated_at.strftime('%Y-%m-%d %H:%M:%S') if auth.updated_at else ''
            ]
            ws.append(row)
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                if col_idx in [1, 3, 6, 7, 9, 10]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"Inspection_Authority_Master_Export_{current_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        from authentication.utils import log_action
        log_action(
            user=request.user,
            action="EXPORT",
            target="All Inspection Authorities Excel",
            module="Inspection Authority Master"
        )
        
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection Authority Template"
        
        headers = [
            "Inspection Authority ID", "Inspection Authority Name", "Category", 
            "Contact Person", "Applicable Standard Number", "Approval Type", "Status", "Remarks"
        ]
        ws.append(headers)
        
        sample_rows = [
            ["IA-001", "Lloyds Register of Shipping", "Marine", "John Doe", "IEC 61439-1", "Marine Approval", "Active", "Standard marine inspection agency."],
            ["IA-002", "Defense QA Agency", "Defence", "Jane Smith", "", "Type Approval", "Active", "Defense grade certifications."],
            ["IA-003", "Customer Inspection", "Customer", "", "ISO 9001", "Standard QA", "Active", "Client representative inspections."]
        ]
        for row in sample_rows:
            ws.append(row)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inspection_authorities_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        import openpyxl
        from django.db import transaction
        from authentication.system_models import AuditLog
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
            
        skip_duplicates = request.data.get('skip_duplicates', 'true').lower() == 'true'
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            
            id_idx = -1
            name_idx = -1
            cat_idx = -1
            contact_idx = -1
            std_idx = -1
            appr_idx = -1
            status_idx = -1
            rem_idx = -1

            for idx, h in enumerate(headers):
                if h in ['inspection authority id', 'inspection_authority_id', 'authority_id', 'id']:
                    id_idx = idx
                elif h in ['inspection authority name', 'inspection_authority_name', 'name']:
                    name_idx = idx
                elif h in ['category', 'type']:
                    cat_idx = idx
                elif h in ['contact person', 'contact_person', 'contact']:
                    contact_idx = idx
                elif h in ['applicable standard number', 'applicable_standard_number', 'standard number', 'standard']:
                    std_idx = idx
                elif h in ['approval type', 'approval_type', 'approval']:
                    appr_idx = idx
                elif h in ['status', 'active']:
                    status_idx = idx
                elif h in ['remarks', 'notes', 'remark']:
                    rem_idx = idx
                    
            missing = []
            if id_idx == -1: missing.append("Inspection Authority ID")
            if name_idx == -1: missing.append("Inspection Authority Name")
            if cat_idx == -1: missing.append("Category")

            if missing:
                return Response({
                    "error": f"Missing required columns: {', '.join(missing)}. Please use the downloadable template."
                }, status=status.HTTP_400_BAD_REQUEST)

            success_count = 0
            skipped_count = 0
            failure_count = 0
            errors = []
            skipped = []
            successes = []

            allowed_cats = ['Marine', 'Customer', 'QA Agency', 'Internal', 'Defence']

            for row_num, row in enumerate(rows[1:], start=2):
                if not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue
                    
                auth_id = ""
                try:
                    auth_id = str(row[id_idx]).strip() if row[id_idx] is not None else ""
                    auth_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                    cat_val = str(row[cat_idx]).strip() if row[cat_idx] is not None else ""
                    
                    if not auth_id:
                        errors.append({
                            "row": row_num,
                            "authority_id": "N/A",
                            "error_message": "Inspection Authority ID is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not auth_name:
                        errors.append({
                            "row": row_num,
                            "authority_id": auth_id,
                            "error_message": "Inspection Authority Name is required"
                        })
                        failure_count += 1
                        continue
                        
                    if not cat_val:
                        errors.append({
                            "row": row_num,
                            "authority_id": auth_id,
                            "error_message": "Category is required"
                        })
                        failure_count += 1
                        continue

                    # Validate category choice
                    match_cat = next((c for c in allowed_cats if c.lower() == cat_val.lower()), None)
                    if not match_cat:
                        errors.append({
                            "row": row_num,
                            "authority_id": auth_id,
                            "error_message": f"Invalid category '{cat_val}'. Must be one of: {', '.join(allowed_cats)}"
                        })
                        failure_count += 1
                        continue

                    contact_val = str(row[contact_idx]).strip() if (contact_idx != -1 and row[contact_idx] is not None) else ""
                    appr_val = str(row[appr_idx]).strip() if (appr_idx != -1 and row[appr_idx] is not None) else ""
                    rem_val = str(row[rem_idx]).strip() if (rem_idx != -1 and row[rem_idx] is not None) else ""
                    
                    # Resolve applicable standard relation if provided
                    std_instance = None
                    if std_idx != -1 and row[std_idx] is not None:
                        std_num = str(row[std_idx]).strip()
                        if std_num:
                            std_instance = StandardMaster.objects.filter(standard_number__iexact=std_num).first()
                            if not std_instance:
                                errors.append({
                                    "row": row_num,
                                    "authority_id": auth_id,
                                    "error_message": f"Standard with number '{std_num}' does not exist. Please upload the Standard first."
                                })
                                failure_count += 1
                                continue
                    
                    status_val = "Active"
                    if status_idx != -1 and row[status_idx] is not None:
                        raw_status = str(row[status_idx]).strip().capitalize()
                        if raw_status in ["Active", "Inactive"]:
                            status_val = raw_status

                    if skip_duplicates:
                        existing = InspectionAuthorityMaster.objects.filter(authority_id__iexact=auth_id).first()
                        if existing:
                            skipped.append({
                                "row": row_num,
                                "authority_id": auth_id,
                                "reason": f"Inspection Authority with ID '{auth_id}' already exists."
                            })
                            skipped_count += 1
                            continue
                            
                    with transaction.atomic():
                        auth_instance, created = InspectionAuthorityMaster.objects.update_or_create(
                            authority_id=auth_id,
                            defaults={
                                "name": auth_name,
                                "category": match_cat,
                                "contact_person": contact_val,
                                "applicable_standard": std_instance,
                                "approval_type": appr_val,
                                "status": status_val,
                                "remarks": rem_val
                            }
                        )
                        
                        AuditLog.objects.create(
                            user=request.user,
                            action="Inspection Authority Created/Updated via Bulk Upload",
                            target=f"Authority: {auth_instance.authority_id}",
                            module="Inspection Authority Master",
                            status="SUCCESS"
                        )
                        
                    successes.append(auth_id)
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": row_num,
                        "authority_id": auth_id if auth_id else "Unknown",
                        "error_message": f"Unexpected error: {str(e)}"
                    })
                    failure_count += 1
                    
            return Response({
                "success_count": success_count,
                "skipped_count": skipped_count,
                "failure_count": failure_count,
                "total_processed": len(rows) - 1,
                "errors": errors,
                "skipped": skipped,
                "successes": successes
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        from django.db.models import Count, Q
        from django.db.models.functions import TruncMonth
        from authentication.system_models import SystemConfiguration
        
        user = request.user
        
        # All users see all projects for full transparency
        projects_qs = Project.objects.all()
        
        # Apply Year and Month Filters
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        
        if year and year != 'all' and year.isdigit():
            projects_qs = projects_qs.filter(date_received__year=int(year))
        if month and month != 'all' and month.isdigit():
            projects_qs = projects_qs.filter(date_received__month=int(month))
            
        # 1. Quick Stats
        total_projects = projects_qs.count()
        closed_projects = projects_qs.filter(status='Closed').count()
        open_projects = projects_qs.filter(status__in=['Open', 'In Progress']).count()
        pending_approval = projects_qs.filter(status='Pending Approval').count()
        customers_count = CustomerMaster.objects.count()
        
        completion_rate = (closed_projects / total_projects * 100) if total_projects > 0 else 0
        
        # 2. Project Type Distribution (Pie Chart)
        type_distribution = projects_qs.values('project_type').annotate(count=Count('id')).order_by('count')
        
        # 2.5 Stage-wise Distribution for Open Projects
        from workflow.models import StageInstance, StageTemplate
        active_projects_ids = projects_qs.filter(status__in=['Open', 'In Progress']).values_list('id', flat=True)
        
        # Get count of projects at each stage (where that stage is the current one)
        # Current stage is defined as the first 'Unlocked' or 'In Progress' stage
        stage_distribution = []
        templates = StageTemplate.objects.filter(is_active=True).order_by('order')
        
        for template in templates:
            # Count projects where this template is their first non-approved stage
            count = 0
            # This is slightly complex in SQL, so we'll approximate: 
            # Count projects where this stage instance is UNLOCKED or SUBMITTED
            count = StageInstance.objects.filter(
                template=template,
                project_id__in=active_projects_ids,
                status__in=['Unlocked', 'In Progress', 'Submitted', 'Rejected']
            ).count()
            
            if count > 0:
                stage_distribution.append({
                    'name': template.name,
                    'count': count
                })
                
        # 3. Monthly Trend (Bar/Line Chart)
        monthly_trend = projects_qs.annotate(
            month=TruncMonth('date_received')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
        
        # 4. Recent Projects (Table)
        recent_projects = projects_qs.select_related('created_by').order_by('-created_at')[:10]
        recent_data = ProjectSerializer(recent_projects, many=True).data
        
        # 5. System Info (Only for Admins/Supervisors, or limited for employees)
        sys_config = SystemConfiguration.objects.first()
        if not sys_config:
            sys_config = SystemConfiguration.objects.create()
            
        dashboard_data = {
            'stats': {
                'total': total_projects,
                'closed': closed_projects,
                'open': open_projects,
                'pending': pending_approval,
                'customers': customers_count,
                'completion_rate': round(completion_rate, 2)
            },
            'charts': {
                'type_distribution': list(type_distribution),
                'stage_distribution': stage_distribution,
                'monthly_trend': [
                    {'month': item['month'].strftime('%b %Y'), 'count': item['count']} 
                    for item in monthly_trend if item['month']
                ]
            },
            'recent_projects': recent_data,
        }
        
        # Add system info for Admins/Supervisors
        if user.role in ['ADMIN', 'SUPERVISOR']:
            dashboard_data['system_info'] = {
                'company': sys_config.company_name,
                'financial_year': sys_config.financial_year,
                'version': sys_config.system_version,
                'last_update': sys_config.last_update,
                'server': sys_config.server_info
            }
        else:
            dashboard_data['system_info'] = {
                'company': sys_config.company_name,
                'financial_year': sys_config.financial_year,
                'version': sys_config.system_version,
                'last_update': sys_config.last_update,
            }
            
        return Response(dashboard_data)


class ECNViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = ECN.objects.all().select_related('project', 'project__customer', 'initiator', 'reviewed_by', 'approved_by')
    serializer_class = ECNSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = ['status', 'project']
    search_fields = ['ecn_number', 'project__name', 'project__customer__name', 'project__customer_name', 'raised_department']
    ordering_fields = ['ecn_date', 'created_at', 'ecn_number']
    ordering = ['-created_at']
    audit_module = "ECN"

    def get_audit_target(self, instance):
        return f"ECN: {instance.ecn_number} for Project: {instance.project.name}"

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAdmin()]
        return [permissions.IsAuthenticated()]

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        
        # Filter by customer name (in case it is custom customer name string or fk relation name)
        customer_name = self.request.query_params.get('customer_name', None)
        if customer_name:
            queryset = queryset.filter(
                Q(project__customer__name__icontains=customer_name) | 
                Q(project__customer_name__icontains=customer_name)
            )
            
        # Filter by project name
        project_name = self.request.query_params.get('project_name', None)
        if project_name:
            queryset = queryset.filter(project__name__icontains=project_name)

        # Filter by date
        date_val = self.request.query_params.get('date', None)
        if date_val:
            queryset = queryset.filter(ecn_date=date_val)
            
        return queryset
